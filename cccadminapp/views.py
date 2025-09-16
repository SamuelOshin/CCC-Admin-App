from django.shortcuts import render
from django.http import HttpResponse
from django.template.loader import get_template
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
from django.db.models import Count, Q, Max, Avg
from django.utils import timezone
from datetime import timedelta, date
import json
import logging
import csv
import io
import zipfile
from collections import defaultdict

# Import models from all apps
from clergy_registration.models import ClergyDetails, AnnointmentGazzette
from ParishRestructure.models import ParishDirectory, ParishRegistration
from transfer.models import TransferData, PostingHistory, ClergyTrfbio

# Import utilities
from .utils import get_month_format

logger = logging.getLogger(__name__)

def landing_page(request):
    """
    Landing page for the CCC Administrative Management System.
    Provides an overview of the application and links to sign-in.
    """
    return render(request, 'landing.html')

@login_required
def centralized_dashboard(request):
    """
    Centralized dashboard that aggregates data from all modules based on user permissions.
    Displays comprehensive statistics and charts for clergy, parishes, transfers, and annointments.
    """
    context = {
        'page_title': 'CCC Administrative Dashboard',
        'page_subtitle': 'Comprehensive overview of all church administrative data',
    }

    try:
        # Get user permissions
        user_groups = list(request.user.groups.values_list('name', flat=True))
        is_superuser = request.user.is_superuser

        # Initialize permission flags
        can_manage_clergy = 'clergyadmin' in user_groups or is_superuser
        can_manage_transfers = 'transferadmin' in user_groups or is_superuser
        can_manage_parishes = 'parishadmin' in user_groups or is_superuser

        context.update({
            'can_manage_clergy': can_manage_clergy,
            'can_manage_transfers': can_manage_transfers,
            'can_manage_parishes': can_manage_parishes,
            'is_superuser': is_superuser,
        })

        # Initialize common date variables used across all sections
        thirty_days_ago = timezone.now().date() - timedelta(days=30)
        six_months_ago = timezone.now().date() - timedelta(days=180)
        current_year = timezone.now().year

        # ===== CLERGY STATISTICS =====
        if can_manage_clergy:
            # Basic clergy counts
            total_clergy = ClergyDetails.objects.count()

            # Recent registrations (last 30 days)
            recent_clergy_registrations = ClergyDetails.objects.filter(
                # Using clergy_id as proxy for recent registrations since no created_at field
                clergy_id__in=ClergyDetails.objects.order_by('-clergy_id')[:20].values_list('clergy_id', flat=True)
            ).count()

            # Gender distribution
            gender_stats = ClergyDetails.objects.values('gender').annotate(
                count=Count('gender')
            ).exclude(gender='').order_by('gender')

            # Marital status distribution
            marital_stats = ClergyDetails.objects.values('marital_status').annotate(
                count=Count('marital_status')
            ).exclude(marital_status='').order_by('marital_status')

            # Education level distribution
            education_stats = ClergyDetails.objects.values('edu_level').annotate(
                count=Count('edu_level')
            ).exclude(edu_level='').order_by('-count')[:5]

            # Rank distribution from AnnointmentGazzette
            rank_stats = AnnointmentGazzette.objects.values('rank').annotate(
                count=Count('rank')
            ).exclude(rank='').order_by('-count')[:8]

            # Recent clergy for table display
            recent_clergy = ClergyDetails.objects.order_by('-clergy_id')[:5]

            # Add rank information to recent clergy
            for clergy in recent_clergy:
                latest_annointment = AnnointmentGazzette.objects.filter(
                    clergy=clergy
                ).order_by('-year_of_annointment', '-month_of_annointment').first()
                clergy.rank = latest_annointment.rank if latest_annointment else 'Not Set'

            context.update({
                'total_clergy': total_clergy,
                'recent_clergy_registrations': recent_clergy_registrations,
                'gender_stats': list(gender_stats),
                'marital_stats': list(marital_stats),
                'education_stats': list(education_stats),
                'rank_stats': list(rank_stats),
                'recent_clergy': recent_clergy,
            })

        # ===== TRANSFER STATISTICS =====
        if can_manage_transfers:
            # Transfer counts by status
            total_transfers = TransferData.objects.count()
            pending_transfers = TransferData.objects.filter(trf_status='Pending').count()
            approved_transfers = TransferData.objects.filter(trf_status='Approved').count()
            withdrawn_transfers = TransferData.objects.filter(trf_status='Withdrawn').count()

            # Floating clergy count
            floating_clergy = ClergyTrfbio.objects.filter(floating=True).count()

            # Recent transfers (last 30 days)
            recent_transfers = TransferData.objects.filter(
                date_transfered__gte=thirty_days_ago
            ).select_related('clergy', 'parishFrm', 'parishTo').order_by('-date_transfered')[:5]

            # Monthly transfer trends (last 6 months)
            monthly_transfers = TransferData.objects.filter(
                date_transfered__gte=six_months_ago
            ).extra(
                select=get_month_format('date_transfered')
            ).values('month').annotate(
                count=Count('id')
            ).order_by('month')

            context.update({
                'total_transfers': total_transfers,
                'pending_transfers': pending_transfers,
                'approved_transfers': approved_transfers,
                'withdrawn_transfers': withdrawn_transfers,
                'floating_clergy': floating_clergy,
                'recent_transfers': recent_transfers,
                'monthly_transfers': list(monthly_transfers),
            })

        # ===== PARISH STATISTICS =====
        if can_manage_parishes:
            # Basic parish counts
            total_parishes = ParishDirectory.objects.count()
            registered_parishes = ParishDirectory.objects.filter(register_status=True).count()
            unregistered_parishes = total_parishes - registered_parishes

            # Registration statistics
            pending_registrations = ParishRegistration.objects.filter(date_approved__isnull=True).count()
            approved_registrations = ParishRegistration.objects.filter(date_approved__isnull=False).count()

            # Recent registrations (last 30 days)
            recent_parish_registrations = ParishRegistration.objects.filter(
                date_applied__gte=thirty_days_ago
            ).count()

            # Monthly registration trends (last 6 months)
            monthly_registrations = ParishRegistration.objects.filter(
                date_applied__gte=six_months_ago
            ).extra(
                select=get_month_format('date_applied')
            ).values('month').annotate(
                count=Count('id')
            ).order_by('month')

            # Diocese distribution
            diocese_stats = ParishRegistration.objects.values('diocese__name').annotate(
                count=Count('diocese')
            ).exclude(diocese__name='').order_by('-count')[:8]

            # Document completion status
            document_fields = [
                'application_for_registration', 'original_receipt_of_land',
                'original_survey_plan', 'building_plan', 'sworn_affidavit',
                'passport_photograph', 'approval_from_government_diaspora',
                'payment_proof_of_auditorium'
            ]

            document_completion = {}
            for field in document_fields:
                completed = ParishRegistration.objects.filter(**{field: True}).count()
                total = ParishRegistration.objects.count()
                percentage = (completed / total * 100) if total > 0 else 0
                document_completion[field.replace('_', ' ').title()] = {
                    'completed': completed,
                    'percentage': round(percentage, 1)
                }

            # Recent parish registrations
            recent_parish_regs = ParishRegistration.objects.select_related(
                'parish', 'diocese'
            ).order_by('-date_applied')[:5]

            context.update({
                'total_parishes': total_parishes,
                'registered_parishes': registered_parishes,
                'unregistered_parishes': unregistered_parishes,
                'pending_registrations': pending_registrations,
                'approved_registrations': approved_registrations,
                'recent_parish_registrations': recent_parish_registrations,
                'monthly_registrations': list(monthly_registrations),
                'diocese_stats': list(diocese_stats),
                'document_completion': document_completion,
                'recent_parish_regs': recent_parish_regs,
            })

        # ===== ANNOINTMENT STATISTICS =====
        if can_manage_clergy:
            # Annointment statistics
            total_annointments = AnnointmentGazzette.objects.count()

            # Annointment places distribution
            place_stats = AnnointmentGazzette.objects.values('place_of_annoitment').annotate(
                count=Count('place_of_annoitment')
            ).exclude(place_of_annoitment='').order_by('-count')[:5]

            # Yearly annointment trends (last 5 years)
            yearly_annointments = []
            for year in range(current_year - 4, current_year + 1):
                count = AnnointmentGazzette.objects.filter(year_of_annointment=year).count()
                yearly_annointments.append({'year': year, 'count': count})

            # Recent annointments
            recent_annointments = AnnointmentGazzette.objects.select_related(
                'clergy'
            ).order_by('-year_of_annointment', '-month_of_annointment')[:5]

            context.update({
                'total_annointments': total_annointments,
                'place_stats': list(place_stats),
                'yearly_annointments': yearly_annointments,
                'recent_annointments': recent_annointments,
            })

        # ===== CROSS-MODULE STATISTICS =====
        # Overall system health metrics
        total_active_entities = 0
        if can_manage_clergy:
            total_active_entities += total_clergy
        if can_manage_parishes:
            total_active_entities += total_parishes
        if can_manage_transfers:
            total_active_entities += total_transfers

        # Recent activity summary (last 7 days)
        seven_days_ago = timezone.now().date() - timedelta(days=7)

        recent_activity = {
            'clergy_registrations': 0,
            'parish_registrations': 0,
            'transfers': 0,
            'annointments': 0,
        }

        if can_manage_clergy:
            recent_activity['clergy_registrations'] = ClergyDetails.objects.filter(
                clergy_id__in=ClergyDetails.objects.order_by('-clergy_id')[:10].values_list('clergy_id', flat=True)
            ).count()
            recent_activity['annointments'] = AnnointmentGazzette.objects.filter(
                # Using year/month as proxy for recent activity
                year_of_annointment__gte=current_year - 1
            ).count()

        if can_manage_parishes:
            recent_activity['parish_registrations'] = ParishRegistration.objects.filter(
                date_applied__gte=seven_days_ago
            ).count()

        if can_manage_transfers:
            recent_activity['transfers'] = TransferData.objects.filter(
                date_transfered__gte=seven_days_ago
            ).count()

        context.update({
            'total_active_entities': total_active_entities,
            'recent_activity': recent_activity,
        })

    except Exception as e:
        logger.error(f"Error generating centralized dashboard data: {str(e)}", exc_info=True)
        context['error_message'] = "Unable to load dashboard data. Please try again later."

    return render(request, 'dashboard/centralized_dashboard.html', context)


@login_required
def analytics_dashboard(request):
    """
    Advanced analytics dashboard with comprehensive data analysis,
    trends, comparisons, and export capabilities.
    """
    context = {
        'page_title': 'Analytics Dashboard',
        'page_subtitle': 'Advanced data analysis and insights',
    }

    try:
        # Get user permissions
        user_groups = list(request.user.groups.values_list('name', flat=True))
        is_superuser = request.user.is_superuser

        # Initialize permission flags
        can_manage_clergy = 'clergyadmin' in user_groups or is_superuser
        can_manage_transfers = 'transferadmin' in user_groups or is_superuser
        can_manage_parishes = 'parishadmin' in user_groups or is_superuser

        context.update({
            'can_manage_clergy': can_manage_clergy,
            'can_manage_transfers': can_manage_transfers,
            'can_manage_parishes': can_manage_parishes,
            'is_superuser': is_superuser,
        })

        # Debug logging
        print(f"DEBUG: User groups: {user_groups}")
        print(f"DEBUG: Is superuser: {is_superuser}")
        print(f"DEBUG: Permissions - clergy: {can_manage_clergy}, transfers: {can_manage_transfers}, parishes: {can_manage_parishes}")

        # ===== ADVANCED ANALYTICS DATA =====

        # Time-based analytics (last 12 months)
        twelve_months_ago = timezone.now().date() - timedelta(days=365)

        # Monthly trends data
        monthly_data = []

        for i in range(12):
            month_date = timezone.now().date() - timedelta(days=30 * (11 - i))
            month_start = month_date.replace(day=1)
            month_end = (month_start + timedelta(days=32)).replace(day=1) - timedelta(days=1)

            month_data = {
                'month': month_start.strftime('%b %Y'),
                'clergy_registrations': 0,
                'transfers': 0,
                'parish_registrations': 0,
                'annointments': 0,
            }

            if can_manage_clergy:
                # Use entry_date_in_ccc for clergy registration date
                month_data['clergy_registrations'] = ClergyDetails.objects.filter(
                    entry_date_in_ccc__year=month_date.year,
                    entry_date_in_ccc__month=month_date.month
                ).count()

                # Use year_of_annointment and month_of_annointment for annointments
                month_data['annointments'] = AnnointmentGazzette.objects.filter(
                    year_of_annointment=month_date.year,
                    month_of_annointment__iexact=month_date.strftime('%B')  # Match month name
                ).count()

            if can_manage_transfers:
                month_data['transfers'] = TransferData.objects.filter(
                    date_transfered__year=month_date.year,
                    date_transfered__month=month_date.month
                ).count()

            if can_manage_parishes:
                month_data['parish_registrations'] = ParishRegistration.objects.filter(
                    date_applied__year=month_date.year,
                    date_applied__month=month_date.month
                ).count()

            monthly_data.append(month_data)

        # ===== PERFORMANCE METRICS =====

        # Growth rates
        current_month = timezone.now().date().replace(day=1)
        last_month = (current_month - timedelta(days=1)).replace(day=1)

        growth_metrics = {}

        if can_manage_clergy:
            current_clergy = ClergyDetails.objects.count()
            last_month_clergy = ClergyDetails.objects.filter(
                entry_date_in_ccc__lt=current_month
            ).count()
            growth_metrics['clergy_growth'] = {
                'current': current_clergy,
                'previous': last_month_clergy,
                'growth_rate': ((current_clergy - last_month_clergy) / max(last_month_clergy, 1)) * 100,
            }

        if can_manage_transfers:
            current_transfers = TransferData.objects.count()
            last_month_transfers = TransferData.objects.filter(
                date_transfered__lt=current_month
            ).count()
            growth_metrics['transfer_growth'] = {
                'current': current_transfers,
                'previous': last_month_transfers,
                'growth_rate': ((current_transfers - last_month_transfers) / max(last_month_transfers, 1)) * 100,
            }

        if can_manage_parishes:
            current_parishes = ParishDirectory.objects.count()
            last_month_parishes = ParishRegistration.objects.filter(
                date_applied__lt=current_month
            ).count()
            growth_metrics['parish_growth'] = {
                'current': current_parishes,
                'previous': last_month_parishes,
                'growth_rate': ((current_parishes - last_month_parishes) / max(last_month_parishes, 1)) * 100,
            }

        # ===== DEMOGRAPHIC ANALYSIS =====

        demographic_data = {}

        if can_manage_clergy:
            # Age distribution (calculate from date of birth)
            from datetime import date
            current_year = date.today().year

            age_distribution = {
                '18-25': ClergyDetails.objects.filter(
                    dob__year__lte=current_year-18,
                    dob__year__gte=current_year-25
                ).count(),
                '26-35': ClergyDetails.objects.filter(
                    dob__year__lte=current_year-26,
                    dob__year__gte=current_year-35
                ).count(),
                '36-45': ClergyDetails.objects.filter(
                    dob__year__lte=current_year-36,
                    dob__year__gte=current_year-45
                ).count(),
                '46-55': ClergyDetails.objects.filter(
                    dob__year__lte=current_year-46,
                    dob__year__gte=current_year-55
                ).count(),
                '56+': ClergyDetails.objects.filter(
                    dob__year__lte=current_year-56
                ).count(),
            }
            demographic_data['age_distribution'] = age_distribution

            # Education distribution (remove Avg aggregation for SQLite compatibility)
            education_performance = ClergyDetails.objects.values('edu_level').annotate(
                count=Count('edu_level')
            ).exclude(edu_level='').order_by('-count')[:5]
            demographic_data['education_performance'] = list(education_performance)

        # ===== GEOGRAPHIC ANALYSIS =====

        geographic_data = {}

        if can_manage_parishes:
            # Parish distribution by diocese (correct field name)
            parish_geographic = ParishRegistration.objects.values('diocese__name').annotate(
                count=Count('diocese')
            ).exclude(diocese__name='').order_by('-count')[:10]
            geographic_data['parish_distribution'] = list(parish_geographic)

        if can_manage_clergy:
            # Clergy distribution by annointment place (correct field name)
            clergy_geographic = AnnointmentGazzette.objects.values('place_of_annoitment').annotate(
                count=Count('place_of_annoitment')
            ).exclude(place_of_annoitment='').order_by('-count')[:10]
            geographic_data['clergy_distribution'] = list(clergy_geographic)

        # ===== EFFICIENCY METRICS =====

        efficiency_metrics = {}

        if can_manage_transfers:
            # Transfer processing time (calculate from actual data)
            total_transfers = TransferData.objects.count()
            pending_transfers = TransferData.objects.filter(trf_status='Pending').count()
            approved_transfers = TransferData.objects.filter(trf_status='Approved').count()
            rejected_transfers = TransferData.objects.filter(trf_status='Withdrawn').count()

            efficiency_metrics['transfer_stats'] = {
                'total': total_transfers,
                'pending': pending_transfers,
                'approved': approved_transfers,
                'rejected': rejected_transfers,
            }

        if can_manage_parishes:
            # Parish registration approval time
            total_registrations = ParishRegistration.objects.count()
            pending_approvals = ParishRegistration.objects.filter(date_approved__isnull=True).count()
            approved_registrations = ParishRegistration.objects.filter(date_approved__isnull=False).count()

            efficiency_metrics['parish_approval_rate'] = {
                'total': total_registrations,
                'pending': pending_approvals,
                'approved': approved_registrations,
                'rate': ((approved_registrations / max(total_registrations, 1)) * 100),
            }

        # ===== PREDICTIVE ANALYTICS =====

        predictive_data = {}

        # Simple trend analysis
        if len(monthly_data) >= 3:
            recent_trend = monthly_data[-3:]
            if can_manage_clergy:
                clergy_trend = [item['clergy_registrations'] for item in recent_trend]
                predictive_data['clergy_trend'] = {
                    'data': clergy_trend,
                    'prediction': sum(clergy_trend) / len(clergy_trend),  # Simple average
                }

        # ===== EXPORT DATA PREPARATION =====

        export_formats = ['CSV', 'Excel', 'PDF', 'JSON']
        available_reports = [
            'Clergy Demographics',
            'Transfer Analytics',
            'Parish Statistics',
            'Monthly Trends',
            'Geographic Distribution',
            'Performance Metrics',
        ]

        # Format data for template consumption
        analytics_data = {
            'total_clergy': ClergyDetails.objects.count() if can_manage_clergy else 0,
            'total_parishes': ParishDirectory.objects.count() if can_manage_parishes else 0,
            'pending_transfers': TransferData.objects.filter(trf_status='Pending').count() if can_manage_transfers else 0,
            'system_efficiency': efficiency_metrics.get('parish_approval_rate', {}).get('rate', 0) if can_manage_parishes else 0,
            'clergy_growth_rate': growth_metrics.get('clergy_growth', {}).get('growth_rate', 0) if can_manage_clergy else 0,
            'parish_growth_rate': growth_metrics.get('parish_growth', {}).get('growth_rate', 0) if can_manage_parishes else 0,
            'avg_transfer_time': 7,  # Could be calculated from actual transfer processing times
            'efficiency_trend': 5,  # Could be calculated from historical efficiency data
            'monthly_trends': {
                'labels': [item['month'] for item in monthly_data],
                'clergy_data': [item['clergy_registrations'] for item in monthly_data],
                'parish_data': [item['parish_registrations'] for item in monthly_data],
            },
            'geographic_data': {
                'labels': [item['diocese__name'] for item in geographic_data.get('parish_distribution', [])],
                'values': [item['count'] for item in geographic_data.get('parish_distribution', [])],
            },
            'demographic_data': {
                'labels': list(demographic_data.get('age_distribution', {}).keys()),
                'values': list(demographic_data.get('age_distribution', {}).values()),
            },
            'transfer_status': {
                'labels': ['Approved', 'Pending', 'Rejected'],
                'values': [
                    efficiency_metrics.get('transfer_stats', {}).get('approved', 0),
                    efficiency_metrics.get('transfer_stats', {}).get('pending', 0),
                    efficiency_metrics.get('transfer_stats', {}).get('rejected', 0),
                ],
            },
            'predictive_data': {
                'labels': ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun'],
                'clergy_projection': [120, 125, 130, 135, 140, 145],  # Could implement actual prediction logic
                'parish_projection': [45, 47, 49, 51, 53, 55],  # Could implement actual prediction logic
            },
            'top_parishes': [
                {'name': 'CCC WorldWide HQtrs ', 'location': 'Imeko, Ogun State', 'clergy_count': '5k+', 'activity_score': 85},
                {'name': 'Nigeria Arch Diocece', 'location': 'Nigeria', 'clergy_count': '3k+', 'activity_score': 78},
                {'name': 'International HQtrs', 'location': 'Lagos,Nigeria', 'clergy_count': '2k+', 'activity_score': 92},
            ],  # Could be calculated from actual parish data
            'recent_activities': [
                {'description': 'New clergy registration', 'user': 'System', 'timestamp': timezone.now().strftime('%Y-%m-%d %H:%M'), 'status': 'completed'},
                {'description': 'Parish transfer request', 'user': 'System', 'timestamp': timezone.now().strftime('%Y-%m-%d %H:%M'), 'status': 'pending'},
                {'description': 'Monthly report generated', 'user': 'System', 'timestamp': timezone.now().strftime('%Y-%m-%d %H:%M'), 'status': 'completed'},
            ],  # Could be populated from actual activity logs
        }

        context.update({
            'monthly_data': monthly_data,
            'growth_metrics': growth_metrics,
            'demographic_data': demographic_data,
            'geographic_data': geographic_data,
            'efficiency_metrics': efficiency_metrics,
            'predictive_data': predictive_data,
            'export_formats': export_formats,
            'available_reports': available_reports,
            'analytics': analytics_data,  # Add formatted analytics data
        })

    except Exception as e:
        logger.error(f"Error generating analytics dashboard data: {str(e)}", exc_info=True)
        context['error_message'] = "Unable to load analytics data. Please try again later."

    return render(request, 'dashboard/analytics_dashboard.html', context)


@login_required
def export_analytics_data(request):
    """
    Export analytics data in various formats (CSV, Excel, PDF, JSON)
    """
    export_format = request.GET.get('format', 'csv')
    report_type = request.GET.get('report', 'all')

    # Get user permissions
    user_groups = list(request.user.groups.values_list('name', flat=True))
    is_superuser = request.user.is_superuser

    can_manage_clergy = 'clergyadmin' in user_groups or is_superuser
    can_manage_transfers = 'transferadmin' in user_groups or is_superuser
    can_manage_parishes = 'parishadmin' in user_groups or is_superuser

    try:
        if export_format == 'csv':
            return export_csv(request, report_type, can_manage_clergy, can_manage_transfers, can_manage_parishes)
        elif export_format == 'json':
            return export_json(request, report_type, can_manage_clergy, can_manage_transfers, can_manage_parishes)
        else:
            return HttpResponse("Unsupported export format. Supported formats: csv, json", status=400)
    except Exception as e:
        logger.error(f"Error exporting analytics data: {str(e)}", exc_info=True)
        return HttpResponse("Error generating export file", status=500)


def export_csv(request, report_type, can_manage_clergy, can_manage_transfers, can_manage_parishes):
    """Export data as CSV"""
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="ccc_analytics_{report_type}_{timezone.now().date()}.csv"'

    writer = csv.writer(response)

    if report_type == 'clergy' and can_manage_clergy:
        writer.writerow(['Clergy ID', 'Name', 'Age', 'Education Level', 'Status'])
        clergy_data = ClergyDetails.objects.all().values_list('clergy_id', 'name', 'age', 'edu_level', 'status')
        for row in clergy_data:
            writer.writerow(row)

    elif report_type == 'transfers' and can_manage_transfers:
        writer.writerow(['Transfer ID', 'Clergy Name', 'From Parish', 'To Parish', 'Date', 'Status'])
        transfer_data = TransferData.objects.all().values_list(
            'id', 'clergy__name', 'from_parish__name', 'to_parish__name', 'date_transfered', 'status'
        )
        for row in transfer_data:
            writer.writerow(row)

    elif report_type == 'parishes' and can_manage_parishes:
        writer.writerow(['Parish Name', 'Location', 'Diocese', 'Status', 'Registration Date'])
        parish_data = ParishRegistration.objects.all().values_list(
            'parish_name', 'location', 'diocese__name', 'status', 'date_applied'
        )
        for row in parish_data:
            writer.writerow(row)

    return response


def export_json(request, report_type, can_manage_clergy, can_manage_transfers, can_manage_parishes):
    """Export data as JSON"""
    data = {}

    if report_type == 'clergy' and can_manage_clergy:
        data['clergy'] = list(ClergyDetails.objects.values(
            'clergy_id', 'name', 'age', 'edu_level', 'status'
        ))

    if report_type == 'transfers' and can_manage_transfers:
        data['transfers'] = list(TransferData.objects.values(
            'id', 'clergy__name', 'from_parish__name', 'to_parish__name', 'date_transfered', 'status'
        ))

    if report_type == 'parishes' and can_manage_parishes:
        data['parishes'] = list(ParishRegistration.objects.values(
            'parish_name', 'location', 'diocese__name', 'status', 'date_applied'
        ))

    response = HttpResponse(json.dumps(data, indent=2, default=str), content_type='application/json')
    response['Content-Disposition'] = f'attachment; filename="ccc_analytics_{report_type}_{timezone.now().date()}.json"'
    return response


@login_required
def export_data(request):
    """
    Comprehensive data export view with filtering, search, and multiple format support.
    Allows users to select data types, apply filters, and export in various formats.
    """
    context = {
        'page_title': 'Data Export Center',
        'page_subtitle': 'Export church administrative data with advanced filtering and search',
    }

    try:
        # Get user permissions
        user_groups = list(request.user.groups.values_list('name', flat=True))
        is_superuser = request.user.is_superuser

        # Initialize permission flags
        can_manage_clergy = 'clergyadmin' in user_groups or is_superuser
        can_manage_transfers = 'transferadmin' in user_groups or is_superuser
        can_manage_parishes = 'parishadmin' in user_groups or is_superuser

        context.update({
            'can_manage_clergy': can_manage_clergy,
            'can_manage_transfers': can_manage_transfers,
            'can_manage_parishes': can_manage_parishes,
            'is_superuser': is_superuser,
        })

        # Available data types for export
        data_types = []

        if can_manage_clergy:
            data_types.extend([
                {
                    'id': 'clergy_details',
                    'name': 'Clergy Details',
                    'description': 'Complete clergy information including personal details, education, and status',
                    'icon': 'fas fa-user-graduate',
                    'count': ClergyDetails.objects.count(),
                    'fields': ['clergy_id', 'name', 'age', 'edu_level', 'status', 'entry_date_in_ccc']
                },
                {
                    'id': 'annointments',
                    'name': 'Annointment Records',
                    'description': 'Clergy annointment records with dates and locations',
                    'icon': 'fas fa-crown',
                    'count': AnnointmentGazzette.objects.count(),
                    'fields': ['clergy__name', 'place_of_annoitment', 'year_of_annointment', 'month_of_annointment']
                }
            ])

        if can_manage_parishes:
            data_types.extend([
                {
                    'id': 'parish_directory',
                    'name': 'Parish Directory',
                    'description': 'Complete parish directory with location and contact information',
                    'icon': 'fas fa-church',
                    'count': ParishDirectory.objects.count(),
                    'fields': ['name', 'location', 'diocese__name', 'contact_info']
                },
                {
                    'id': 'parish_registrations',
                    'name': 'Parish Registrations',
                    'description': 'Parish registration records with application details',
                    'icon': 'fas fa-file-signature',
                    'count': ParishRegistration.objects.count(),
                    'fields': ['parish_name', 'location', 'diocese__name', 'status', 'date_applied']
                }
            ])

        if can_manage_transfers:
            data_types.extend([
                {
                    'id': 'transfer_records',
                    'name': 'Transfer Records',
                    'description': 'Clergy transfer records with dates and parish information',
                    'icon': 'fas fa-exchange-alt',
                    'count': TransferData.objects.count(),
                    'fields': ['clergy__name', 'from_parish__name', 'to_parish__name', 'date_transfered', 'trf_status']
                },
                {
                    'id': 'posting_history',
                    'name': 'Posting History',
                    'description': 'Historical record of clergy postings and designations',
                    'icon': 'fas fa-history',
                    'count': PostingHistory.objects.count(),
                    'fields': ['clergy__name', 'parish__name', 'designation', 'date_of_entry', 'date_of_exit']
                },
                {
                    'id': 'clergy_transfer_bio',
                    'name': 'Clergy Transfer Bio',
                    'description': 'Clergy transfer biography and floating status',
                    'icon': 'fas fa-user-clock',
                    'count': ClergyTrfbio.objects.count(),
                    'fields': ['clergy__name', 'floating']
                }
            ])

        # Add system data for superusers
        if is_superuser:
            from django.contrib.auth.models import User
            data_types.extend([
                {
                    'id': 'users',
                    'name': 'System Users',
                    'description': 'User accounts and permissions',
                    'icon': 'fas fa-users-cog',
                    'count': User.objects.count(),
                    'fields': ['username', 'email', 'first_name', 'last_name', 'is_active', 'date_joined']
                }
            ])

        context['data_types'] = data_types

        # Export formats
        context['export_formats'] = [
            {'id': 'csv', 'name': 'CSV', 'description': 'Comma-separated values for spreadsheet applications', 'icon': 'fas fa-file-csv'},
            {'id': 'excel', 'name': 'Excel', 'description': 'Microsoft Excel format with multiple sheets', 'icon': 'fas fa-file-excel'},
            {'id': 'json', 'name': 'JSON', 'description': 'JavaScript Object Notation for developers', 'icon': 'fas fa-file-code'},
            {'id': 'pdf', 'name': 'PDF Report', 'description': 'Formatted PDF report with charts and summaries', 'icon': 'fas fa-file-pdf'},
        ]

        # Handle POST request for export
        if request.method == 'POST':
            return handle_export_request(request, can_manage_clergy, can_manage_transfers, can_manage_parishes, is_superuser)

    except Exception as e:
        logger.error(f"Error in export data view: {str(e)}", exc_info=True)
        context['error_message'] = "Unable to load export data. Please try again later."

    return render(request, 'dashboard/export_data.html', context)


def handle_export_request(request, can_manage_clergy, can_manage_transfers, can_manage_parishes, is_superuser):
    """
    Handle the export request with filtering and formatting.
    """
    try:
        # Get export parameters
        data_type = request.POST.get('data_type')
        export_format = request.POST.get('export_format', 'csv')
        date_from = request.POST.get('date_from')
        date_to = request.POST.get('date_to')
        status_filter = request.POST.get('status_filter')
        search_query = request.POST.get('search_query')
        selected_fields = request.POST.getlist('selected_fields')

        # Validate data type permissions
        if not validate_data_type_permission(data_type, can_manage_clergy, can_manage_transfers, can_manage_parishes, is_superuser):
            return HttpResponse("Access denied", status=403)

        # Build queryset with filters
        queryset = build_filtered_queryset(
            data_type, date_from, date_to, status_filter, search_query,
            can_manage_clergy, can_manage_transfers, can_manage_parishes, is_superuser
        )

        # Export based on format
        if export_format == 'csv':
            return export_as_csv(queryset, data_type, selected_fields)
        elif export_format == 'excel':
            return export_as_excel(queryset, data_type, selected_fields)
        elif export_format == 'json':
            return export_as_json(queryset, data_type, selected_fields)
        elif export_format == 'pdf':
            return export_as_pdf(queryset, data_type, selected_fields)
        else:
            return HttpResponse("Invalid export format", status=400)

    except Exception as e:
        logger.error(f"Error handling export request: {str(e)}", exc_info=True)
        return HttpResponse("Export failed. Please try again.", status=500)


def validate_data_type_permission(data_type, can_manage_clergy, can_manage_transfers, can_manage_parishes, is_superuser):
    """Validate if user has permission to export the selected data type."""
    permission_map = {
        'clergy_details': can_manage_clergy,
        'annointments': can_manage_clergy,
        'parish_directory': can_manage_parishes,
        'parish_registrations': can_manage_parishes,
        'transfer_records': can_manage_transfers,
        'users': is_superuser,
    }

    return permission_map.get(data_type, False)


def build_filtered_queryset(data_type, date_from, date_to, status_filter, search_query,
                          can_manage_clergy, can_manage_transfers, can_manage_parishes, is_superuser):
    """Build queryset with applied filters."""
    queryset = None

    # Date filters
    date_filters = {}
    if date_from:
        date_filters['date__gte'] = date_from
    if date_to:
        date_filters['date__lte'] = date_to

    if data_type == 'clergy_details' and can_manage_clergy:
        queryset = ClergyDetails.objects.all()
        if date_filters:
            queryset = queryset.filter(**{f'entry_date_in_ccc__{k}': v for k, v in date_filters.items()})
        if status_filter:
            queryset = queryset.filter(status__icontains=status_filter)
        if search_query:
            queryset = queryset.filter(
                Q(name__icontains=search_query) |
                Q(clergy_id__icontains=search_query)
            )

    elif data_type == 'annointments' and can_manage_clergy:
        queryset = AnnointmentGazzette.objects.select_related('clergy')
        if date_filters:
            queryset = queryset.filter(**{f'year_of_annointment__{k}': v for k, v in date_filters.items()})
        if search_query:
            queryset = queryset.filter(
                Q(clergy__first_name__icontains=search_query) |
                Q(clergy__last_name__icontains=search_query) |
                Q(place_of_annoitment__icontains=search_query)
            )

    elif data_type == 'parish_directory' and can_manage_parishes:
        queryset = ParishDirectory.objects.all()  # Remove select_related since ParishDirectory doesn't have diocese field
        if search_query:
            queryset = queryset.filter(
                Q(name__icontains=search_query) |
                Q(address__icontains=search_query)
            )

    elif data_type == 'parish_registrations' and can_manage_parishes:
        queryset = ParishRegistration.objects.select_related('parish', 'diocese')
        if date_filters:
            queryset = queryset.filter(**{f'date_applied__{k}': v for k, v in date_filters.items()})
        if status_filter:
            queryset = queryset.filter(register_status__icontains=status_filter)
        if search_query:
            queryset = queryset.filter(
                Q(parish__name__icontains=search_query) |
                Q(country__icontains=search_query) |
                Q(state__icontains=search_query) |
                Q(city__icontains=search_query) |
                Q(diocese__name__icontains=search_query)
            )

    elif data_type == 'transfer_records' and can_manage_transfers:
        queryset = TransferData.objects.select_related('clergy', 'parishFrm', 'parishTo')
        if date_filters:
            queryset = queryset.filter(**{f'date_transfered__{k}': v for k, v in date_filters.items()})
        if status_filter:
            queryset = queryset.filter(trf_status__icontains=status_filter)
        if search_query:
            queryset = queryset.filter(
                Q(clergy__first_name__icontains=search_query) |
                Q(clergy__last_name__icontains=search_query) |
                Q(parishFrm__parish__name__icontains=search_query) |
                Q(parishTo__parish__name__icontains=search_query)
            )

    elif data_type == 'users' and is_superuser:
        from django.contrib.auth.models import User
        queryset = User.objects.all()
        if date_filters:
            queryset = queryset.filter(**{f'date_joined__{k}': v for k, v in date_filters.items()})
        if status_filter:
            is_active = status_filter.lower() == 'active'
            queryset = queryset.filter(is_active=is_active)
        if search_query:
            queryset = queryset.filter(
                Q(username__icontains=search_query) |
                Q(email__icontains=search_query) |
                Q(first_name__icontains=search_query) |
                Q(last_name__icontains=search_query)
            )

    return queryset


def export_as_csv(queryset, data_type, selected_fields):
    """Export data as CSV format."""
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="{data_type}_{timezone.now().date()}.csv"'

    writer = csv.writer(response)

    # Write headers
    if selected_fields:
        writer.writerow(selected_fields)
    else:
        # Default headers based on data type
        headers = get_default_headers(data_type)
        writer.writerow(headers)

    # Write data
    for obj in queryset:
        row = get_object_row(obj, data_type, selected_fields)
        writer.writerow(row)

    return response


def export_as_excel(queryset, data_type, selected_fields):
    """Export data as Excel format."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill
    except ImportError:
        return HttpResponse("Excel export requires openpyxl. Please install it.", status=500)

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{data_type}_{timezone.now().date()}.xlsx"'

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = data_type.replace('_', ' ').title()

    # Headers with styling
    headers = selected_fields if selected_fields else get_default_headers(data_type)
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        # Auto-adjust column width
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = max(len(header) + 2, 15)

    # Data with alternating row colors
    for row_num, obj in enumerate(queryset, 2):
        row_data = get_object_row(obj, data_type, selected_fields)
        fill = PatternFill(start_color="E6F3FF" if row_num % 2 == 0 else "FFFFFF", end_color="E6F3FF" if row_num % 2 == 0 else "FFFFFF", fill_type="solid")

        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num, value=str(value) if value is not None else '')
            cell.fill = fill

    wb.save(response)
    return response


def export_as_json(queryset, data_type, selected_fields):
    """Export data as JSON format."""
    data = []
    for obj in queryset:
        row = get_object_row(obj, data_type, selected_fields)
        headers = selected_fields if selected_fields else get_default_headers(data_type)
        data.append(dict(zip(headers, row)))

    response = HttpResponse(
        json.dumps(data, indent=2, default=str),
        content_type='application/json'
    )
    response['Content-Disposition'] = f'attachment; filename="{data_type}_{timezone.now().date()}.json"'
    return response


def export_as_pdf(queryset, data_type, selected_fields):
    """Export data as PDF format with enhanced formatting."""
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import letter, A4
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import inch
        from io import BytesIO
    except ImportError:
        return HttpResponse("PDF export requires reportlab. Please install with: pip install reportlab", status=500)

    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=72, leftMargin=72, topMargin=72, bottomMargin=72)
    elements = []

    styles = getSampleStyleSheet()

    # Custom styles
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=16,
        spaceAfter=30,
        alignment=1  # Center alignment
    )

    # Title with data type and export info
    title = Paragraph(f"{data_type.replace('_', ' ').title()} Data Export", title_style)
    elements.append(title)

    # Export metadata
    metadata_style = ParagraphStyle(
        'Metadata',
        parent=styles['Normal'],
        fontSize=10,
        textColor=colors.gray,
        alignment=1
    )

    metadata = Paragraph(f"Generated on: {timezone.now().strftime('%Y-%m-%d %H:%M:%S')}<br/>Total Records: {queryset.count()}", metadata_style)
    elements.append(metadata)
    elements.append(Spacer(1, 20))

    # Data table
    headers = selected_fields if selected_fields else get_default_headers(data_type)
    data = [headers]

    # Limit records for PDF to prevent memory issues
    max_records = min(queryset.count(), 1000)
    for obj in queryset[:max_records]:
        row = get_object_row(obj, data_type, selected_fields)
        data.append(row)

    # Calculate column widths based on content
    col_widths = []
    for i, header in enumerate(headers):
        max_width = len(header)
        for row in data[1:][:10]:  # Check first 10 rows for width calculation
            if i < len(row):
                max_width = max(max_width, len(str(row[i] or '')))
        col_widths.append(min(max_width * 6, 80))  # Max width of 80 points

    table = Table(data, colWidths=col_widths)

    # Enhanced table styling
    table.setStyle(TableStyle([
        # Header styling
        ('BACKGROUND', (0, 0), (-1, 0), colors.darkblue),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
        ('TOPPADDING', (0, 0), (-1, 0), 8),

        # Data row styling
        ('BACKGROUND', (0, 1), (-1, -1), colors.white),
        ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('TOPPADDING', (0, 1), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 4),

        # Alternating row colors
        ('BACKGROUND', (0, 1), (-1, 1), colors.lightgrey),
        ('BACKGROUND', (0, 3), (-1, 3), colors.lightgrey),
        ('BACKGROUND', (0, 5), (-1, 5), colors.lightgrey),

        # Grid lines
        ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
        ('LINEBELOW', (0, 0), (-1, 0), 1, colors.black),
    ]))

    elements.append(table)

    # Add record count info if truncated
    if queryset.count() > max_records:
        warning_style = ParagraphStyle(
            'Warning',
            parent=styles['Normal'],
            fontSize=9,
            textColor=colors.red,
            alignment=1
        )
        warning = Paragraph(f"Note: Only showing first {max_records} records. Total available: {queryset.count()}", warning_style)
        elements.append(Spacer(1, 10))
        elements.append(warning)

    doc.build(elements)
    buffer.seek(0)

    response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{data_type}_{timezone.now().date()}.pdf"'
    return response


def bulk_export_data(request):
    """Export multiple data types in a single ZIP file."""
    if request.method != 'POST':
        return HttpResponse("Method not allowed", status=405)

    try:
        # Get selected data types and format
        selected_types = request.POST.getlist('data_types')
        export_format = request.POST.get('format', 'excel')
        selected_fields = request.POST.getlist('selected_fields')

        if not selected_types:
            return HttpResponse("No data types selected", status=400)

        # Get user permissions
        user_groups = list(request.user.groups.values_list('name', flat=True))
        is_superuser = request.user.is_superuser

        # Create ZIP file in memory
        buffer = io.BytesIO()
        with zipfile.ZipFile(buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
            for data_type in selected_types:
                if not has_permission_for_data_type(data_type, user_groups, is_superuser):
                    continue  # Skip data types user doesn't have permission for

                # Get queryset for this data type
                queryset = get_filtered_queryset(data_type, {}, user_groups, is_superuser)

                if queryset.exists():
                    # Generate export based on format
                    if export_format == 'excel':
                        response = export_as_excel(queryset, data_type, selected_fields)
                    elif export_format == 'csv':
                        response = export_as_csv(queryset, data_type, selected_fields)
                    elif export_format == 'json':
                        response = export_as_json(queryset, data_type, selected_fields)
                    elif export_format == 'pdf':
                        response = export_as_pdf(queryset, data_type, selected_fields)
                    else:
                        continue

                    # Add file to ZIP
                    filename = f"{data_type}_{timezone.now().date()}.{get_file_extension(export_format)}"
                    zip_file.writestr(filename, response.content)

        buffer.seek(0)

        # Create response with ZIP file
        response = HttpResponse(buffer.getvalue(), content_type='application/zip')
        response['Content-Disposition'] = f'attachment; filename="bulk_export_{timezone.now().date()}.zip"'
        return response

    except Exception as e:
        logger.error(f"Bulk export error: {str(e)}")
        return HttpResponse(f"Export failed: {str(e)}", status=500)


def get_file_extension(format_type):
    """Get file extension for export format."""
    extensions = {
        'excel': 'xlsx',
        'csv': 'csv',
        'json': 'json',
        'pdf': 'pdf'
    }
    return extensions.get(format_type, 'txt')


def has_permission_for_data_type(data_type, user_groups, is_superuser):
    """Check if user has permission to access a specific data type."""
    if is_superuser:
        return True

    permission_map = {
        'clergy_details': 'clergyadmin' in user_groups,
        'annointments': 'clergyadmin' in user_groups,
        'parish_directory': True,  # All users can access parish directory
        'parish_registrations': True,  # All users can access parish registrations
        'transfer_records': 'transferadmin' in user_groups,
        'users': is_superuser,  # Only superusers can export user data
    }
    return permission_map.get(data_type, False)


def get_filtered_queryset(data_type, filters, user_groups, is_superuser):
    """Get filtered queryset for a data type based on user permissions."""
    if not has_permission_for_data_type(data_type, user_groups, is_superuser):
        return None

    querysets = {
        'clergy_details': ClergyDetails.objects.all(),
        'annointments': AnnointmentGazzette.objects.select_related('clergy'),
        'parish_directory': ParishDirectory.objects.all(),
        'parish_registrations': ParishRegistration.objects.select_related('parish', 'diocese'),
        'transfer_records': TransferData.objects.select_related('clergy', 'parishFrm', 'parishTo'),
        'users': User.objects.all() if is_superuser else User.objects.none(),
    }

    queryset = querysets.get(data_type)
    if queryset is None:
        return None

    # Apply basic filtering if needed
    # This can be extended based on the filters parameter
    return queryset


def get_default_headers(data_type):
    """Get default headers for each data type - ALL FIELDS."""
    headers_map = {
        'clergy_details': [
            'Clergy ID', 'Registration Number', 'Training Number', 'Profile Picture',
            'First Name', 'Middle Name', 'Last Name', 'Alias', 'Gender', 'Marital Status',
            'Date of Birth', 'Spoken Languages', 'Place of Birth', 'Nationality',
            'State of Origin', 'LGA (if Nigerian)', 'Blood Group', 'Genotype',
            'Any Ailment', 'Any Disabilities', 'Ailment Details', 'Disability Details',
            'Permanent Address', 'Resident Address', 'Parish', 'Parish Address',
            'Telephone', 'Email Address', 'Former Religion', 'Denomination',
            'Status in Former Religion', 'Entry Date in CCC', 'First Parish',
            'Shepherd Who Baptized', 'Shepherd Who Sanctified', 'Date When Baptized',
            'Parish Where Baptized', 'First Annointment', 'Date of First Annointment',
            'Present Annointment', 'Date of Present Annointment', 'Education Level',
            'Education Qualification', 'Apprenticeship', 'Hobbies', 'Area of Calling',
            'Working Experience Option', 'Work Experience Details', 'Spouse',
            'Father', 'Mother', 'Next of Kin', 'Relation in CCC', 'Children Info'
        ],
        'annointments': [
            'Clergy Name', 'Place of Annointment', 'Year of Annointment',
            'Month of Annointment', 'Rank', 'Annoiter'
        ],
        'parish_directory': [
            'Parish Name', 'Address', 'Registration Status'
        ],
        'parish_registrations': [
            'Parish Name', 'Country', 'State', 'City', 'Diocese',
            'Date of Establishment', 'Founding Patron', 'Name of Shepherd',
            'Phone', 'Email', 'Date Applied', 'Date Approved',
            'Date Certificate Issued', 'Notes', 'Parish Picture',
            'Application for Registration', 'Original Receipt of Land',
            'Original Survey Plan', 'Building Plan', 'Sworn Affidavit',
            'Passport Photograph', 'Approval from Government/Diaspora',
            'Payment Proof of Auditorium'
        ],
        'transfer_records': [
            'Clergy Name', 'From Parish', 'To Parish', 'Transfer Date',
            'Transfer Begin', 'Transfer End', 'Transfer Status',
            'Designation From', 'Designation To', 'Transfer Extended',
            'Extended Date', 'Remarks', 'Days in Position', 'Days Left'
        ],
        'users': [
            'Username', 'Email', 'First Name', 'Last Name', 'Active',
            'Staff Status', 'Superuser Status', 'Join Date', 'Last Login',
            'Groups', 'User Permissions'
        ],
    }
    return headers_map.get(data_type, [])


def get_object_row(obj, data_type, selected_fields):
    """Get row data for an object based on data type - ALL FIELDS."""
    if data_type == 'clergy_details':
        # Calculate age from date of birth
        age = ''
        if obj.dob:
            from datetime import date
            today = date.today()
            age = today.year - obj.dob.year - ((today.month, today.day) < (obj.dob.month, obj.dob.day))

        return [
            obj.clergy_id,
            obj.reg_number,
            obj.trg_number,
            obj.profile_picture.url if obj.profile_picture else '',
            obj.first_name,
            obj.middle_name,
            obj.last_name,
            obj.alias,
            obj.gender,
            obj.marital_status,
            obj.dob.strftime('%Y-%m-%d') if obj.dob else '',
            ', '.join(obj.spoken_languages) if obj.spoken_languages else '',
            obj.place_of_birth,
            obj.nationality,
            obj.state_of_origin,
            obj.lga_if_nigerian,
            obj.blood_group,
            obj.genotype,
            obj.any_ailment,
            obj.any_disabilities,
            obj.ailment,
            obj.disability,
            obj.permanent_address,
            obj.resident_address,
            obj.parish,
            obj.parish_address,
            str(obj.telephone) if obj.telephone else '',
            obj.email_address,
            obj.former_religion,
            obj.denomination,
            obj.status_former_religion,
            obj.entry_date_in_ccc.strftime('%Y-%m-%d') if obj.entry_date_in_ccc else '',
            obj.first_parish,
            obj.shepherd_who_baptized_you,
            obj.shepherd_who_sanctified_you,
            obj.date_when_baptized.strftime('%Y-%m-%d') if obj.date_when_baptized else '',
            obj.parish_where_baptized,
            obj.first_annointment,
            obj.date_of_first_annointment.strftime('%Y-%m-%d') if obj.date_of_first_annointment else '',
            obj.present_annointment,
            obj.date_of_present_annointment.strftime('%Y-%m-%d') if obj.date_of_present_annointment else '',
            ', '.join(obj.edu_level) if obj.edu_level else '',
            ', '.join(obj.edu_qualification) if obj.edu_qualification else '',
            obj.apprenticeship,
            obj.hobbies,
            ', '.join(obj.area_of_calling) if obj.area_of_calling else '',
            obj.working_experience_option,
            obj.work_experience_ifyes,
            obj.spouse,
            obj.father,
            obj.mother,
            obj.next_of_kin,
            obj.relation_in_ccc,
            obj.children_info,
        ]
    elif data_type == 'annointments':
        return [
            obj.clergy.get_full_name() if obj.clergy else '',
            obj.place_of_annoitment,
            obj.year_of_annointment,
            obj.month_of_annointment,
            obj.rank,
            obj.annoiter,
        ]
    elif data_type == 'parish_directory':
        return [
            obj.name,
            obj.address,
            'Registered' if obj.register_status else 'Not Registered',
        ]
    elif data_type == 'parish_registrations':
        return [
            obj.parish.name if obj.parish else '',
            obj.country,
            obj.state,
            obj.city,
            obj.diocese.name if obj.diocese else '',
            obj.date_of_establishment.strftime('%Y-%m-%d') if obj.date_of_establishment else '',
            obj.founding_patron,
            obj.name_of_shepherd,
            obj.phone,
            obj.email,
            obj.date_applied.strftime('%Y-%m-%d %H:%M:%S') if obj.date_applied else '',
            obj.date_approved.strftime('%Y-%m-%d %H:%M:%S') if obj.date_approved else '',
            obj.date_issued_certificate.strftime('%Y-%m-%d %H:%M:%S') if obj.date_issued_certificate else '',
            obj.notes,
            obj.parish_picture.url if obj.parish_picture else '',
            'Yes' if obj.application_for_registration else 'No',
            'Yes' if obj.original_receipt_of_land else 'No',
            'Yes' if obj.original_survey_plan else 'No',
            'Yes' if obj.building_plan else 'No',
            'Yes' if obj.sworn_affidavit else 'No',
            'Yes' if obj.passport_photograph else 'No',
            'Yes' if obj.approval_from_government_diaspora else 'No',
            'Yes' if obj.payment_proof_of_auditorium else 'No',
        ]
    elif data_type == 'transfer_records':
        return [
            obj.clergy.get_full_name() if obj.clergy else '',
            obj.parishFrm.parish.name if obj.parishFrm and obj.parishFrm.parish else '',
            obj.parishTo.parish.name if obj.parishTo and obj.parishTo.parish else '',
            obj.date_transfered.strftime('%Y-%m-%d') if obj.date_transfered else '',
            obj.trf_begin.strftime('%Y-%m-%d') if obj.trf_begin else '',
            obj.trf_end.strftime('%Y-%m-%d') if obj.trf_end else '',
            obj.trf_status,
            obj.designation_frm,
            obj.designation_to,
            'Yes' if obj.trf_extended else 'No',
            obj.extended_date.strftime('%Y-%m-%d') if obj.extended_date else '',
            obj.remarks,
            str(obj.days_in_position) if hasattr(obj, 'days_in_position') else '',
            str(obj.days_left) if hasattr(obj, 'days_left') and obj.days_left is not None else '',
        ]
    elif data_type == 'users':
        return [
            obj.username,
            obj.email,
            obj.first_name,
            obj.last_name,
            'Yes' if obj.is_active else 'No',
            'Yes' if obj.is_staff else 'No',
            'Yes' if obj.is_superuser else 'No',
            obj.date_joined.strftime('%Y-%m-%d %H:%M:%S') if obj.date_joined else '',
            obj.last_login.strftime('%Y-%m-%d %H:%M:%S') if obj.last_login else '',
            ', '.join([group.name for group in obj.groups.all()]),
            ', '.join([perm.codename for perm in obj.user_permissions.all()]),
        ]
    return []
    return []
